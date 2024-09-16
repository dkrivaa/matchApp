import os
import subprocess
import pandas as pd
import time

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def make_excel_file(couples_list):
    df = pd.DataFrame(couples_list, columns=['file 1', 'file 2'])
    df = df.reset_index(drop=True)
    df.to_excel('match.xlsx', index=False)

    # Format Excel file
    # format_excel_file('simulation.xlsx')

    # Save new EXCEL file with match results
    # Add the Excel file to the Git staging area
    subprocess.run(['git', 'add', 'match.xlsx'], check=True)

    # Commit the Excel file with a message
    commit_message = 'Add Excel file generated by GitHub Actions workflow'
    subprocess.run(['git', 'commit', '-m', commit_message], check=True)

    # Push the changes back to the repository
    subprocess.run(['git', 'push'], check=True)



